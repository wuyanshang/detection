#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
召回评测脚本 (recall-eval)

目的：在安全分类分级功能完整落地前，用标注数据独立量化“召回质量”和“判定质量”，
用真实数据决定优化方向（召回侧 vs 判定侧）。

核心指标：
  - recall@K        正确 category 是否出现在融合后 top-K（衡量召回瓶颈）
  - vector/bm25 各自 recall@K、MRR
  - judge accuracy  在“召回命中”子集上 LLM 选对的比例（衡量判定瓶颈，需 --judge）
  - end2end         端到端准确率（需 --judge）

支持横向对比三种召回策略（验证文档 13.1 方案A 的收益）：
  baseline    仅用字段中文名 columnChnName
  multifield  多字段拼接（表名 + 字段中文名 + 英文名）
  rewrite     LLM 改写为规范检索词（英文缩写展开等，需 LLM）

环境变量：
  ES_HOST            例 http://127.0.0.1:9200
  ES_USERNAME        可选
  ES_PASSWORD        可选
  ES_INDEX           默认 safe_all_topic
  DASHSCOPE_API_KEY  embedding 与 LLM 调用所需
  DASHSCOPE_BASE_URL 默认 https://dashscope.aliyuncs.com/compatible-mode/v1
  EMBED_MODEL        默认 text-embedding-v4
  EMBED_DIM          默认 1024
  LLM_MODEL          默认 qwen-plus

用法示例：
  python recall_eval.py --data labeled_sample.csv --strategy all
  python recall_eval.py --data labeled_sample.xlsx --strategy multifield --judge
  python recall_eval.py --data labeled_sample.csv --k 10 --limit 50 --out report
"""
import argparse
import csv
import json
import os
import sys
import time
from collections import OrderedDict

import requests

# ---------------------------------------------------------------------------
# 配置（环境变量）
# ---------------------------------------------------------------------------
ES_HOST = os.getenv("ES_HOST", "http://127.0.0.1:9200").rstrip("/")
ES_USERNAME = os.getenv("ES_USERNAME") or None
ES_PASSWORD = os.getenv("ES_PASSWORD") or None
ES_INDEX = os.getenv("ES_INDEX", "safe_all_topic")

DASHSCOPE_API_KEY = os.getenv("DASHSCOPE_API_KEY")
DASHSCOPE_BASE_URL = os.getenv(
    "DASHSCOPE_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1"
).rstrip("/")
EMBED_MODEL = os.getenv("EMBED_MODEL", "text-embedding-v4")
EMBED_DIM = int(os.getenv("EMBED_DIM", "1024"))
LLM_MODEL = os.getenv("LLM_MODEL", "qwen-plus")

HTTP_TIMEOUT = int(os.getenv("HTTP_TIMEOUT", "30"))

REQUIRED_FIELDS = ["columnChnName"]
ALL_FIELDS = [
    "systemName",
    "systemDesc",
    "tableName",
    "tableChnName",
    "columnName",
    "columnChnName",
]
GOLD_FIELD = "expected_category"

_es_session = requests.Session()
if ES_USERNAME and ES_PASSWORD:
    _es_session.auth = (ES_USERNAME, ES_PASSWORD)

_embed_cache = {}


# ---------------------------------------------------------------------------
# 数据加载
# ---------------------------------------------------------------------------
def load_data(path):
    """读取标注数据，支持 .csv / .xlsx。每行需含 expected_category 作为金标。"""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        rows = _load_csv(path)
    elif ext in (".xlsx", ".xls"):
        rows = _load_xlsx(path)
    else:
        raise ValueError(f"不支持的文件类型: {ext}（仅支持 .csv / .xlsx）")

    cleaned = []
    for i, row in enumerate(rows):
        if not row.get("columnChnName") and not row.get("columnName"):
            continue
        if not row.get(GOLD_FIELD):
            print(f"  [warn] 第 {i + 1} 行缺少 {GOLD_FIELD}，跳过", file=sys.stderr)
            continue
        cleaned.append(row)
    return cleaned


def _load_csv(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return [dict(r) for r in csv.DictReader(f)]


def _load_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("读取 xlsx 需要 openpyxl：pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    out = []
    for r in rows_iter:
        out.append({header[i]: ("" if v is None else str(v)) for i, v in enumerate(r)})
    return out


# ---------------------------------------------------------------------------
# Embedding
# ---------------------------------------------------------------------------
def embed_text(text):
    text = (text or "").strip()
    if not text:
        raise ValueError("embedding 输入为空")
    if text in _embed_cache:
        return _embed_cache[text]
    if not DASHSCOPE_API_KEY:
        raise SystemExit("缺少 DASHSCOPE_API_KEY，无法生成 embedding")

    payload = {"model": EMBED_MODEL, "input": text, "dimensions": EMBED_DIM}
    headers = {"Authorization": f"Bearer {DASHSCOPE_API_KEY}"}
    resp = requests.post(
        f"{DASHSCOPE_BASE_URL}/embeddings",
        headers=headers,
        json=payload,
        timeout=HTTP_TIMEOUT,
    )
    resp.raise_for_status()
    vec = resp.json()["data"][0]["embedding"]
    _embed_cache[text] = vec
    return vec


# ---------------------------------------------------------------------------
# ES 检索
# ---------------------------------------------------------------------------
def es_vector_search(query_text, top_k, min_score):
    body = {
        "knn": {
            "field": "asset_embedding",
            "query_vector": embed_text(query_text),
            "k": top_k,
            "num_candidates": max(top_k * 10, 100),
        },
        "min_score": min_score,
        "_source": ["asset", "category", "topic"],
        "size": top_k,
    }
    return _es_search(body)


def es_bm25_search(query_text, top_k):
    body = {
        "query": {"match": {"asset": {"query": query_text}}},
        "_source": ["asset", "category", "topic"],
        "size": top_k,
    }
    return _es_search(body)


def _es_search(body):
    resp = _es_session.post(
        f"{ES_HOST}/{ES_INDEX}/_search",
        json=body,
        headers={"Content-Type": "application/json"},
        timeout=HTTP_TIMEOUT,
    )
    resp.raise_for_status()
    hits = resp.json().get("hits", {}).get("hits", [])
    out = []
    for h in hits:
        src = h.get("_source", {})
        out.append(
            {
                "category": src.get("category", ""),
                "asset": src.get("asset", ""),
                "score": h.get("_score", 0.0),
            }
        )
    return out


# ---------------------------------------------------------------------------
# RRF 融合
# ---------------------------------------------------------------------------
def rrf_fuse(*ranked_lists, k=60):
    """对多个按相关性降序排列的候选列表做 Reciprocal Rank Fusion，按 category 去重。"""
    scores = {}
    meta = {}
    for ranked in ranked_lists:
        for rank, item in enumerate(ranked):
            cat = normalize_category(item["category"])
            if not cat:
                continue
            scores[cat] = scores.get(cat, 0.0) + 1.0 / (k + rank + 1)
            if cat not in meta:
                meta[cat] = item
    fused = sorted(scores.items(), key=lambda kv: kv[1], reverse=True)
    return [{"category": c, "rrf": s, "asset": meta[c]["asset"]} for c, s in fused]


# ---------------------------------------------------------------------------
# Query 改写（LLM）
# ---------------------------------------------------------------------------
REWRITE_PROMPT = """你是检索词归一化助手。请根据下面的字段信息，生成一个最适合检索“数据分类目录”的中文规范检索词。
要求：展开英文缩写、补全语义、去掉无意义的单位/编号，只输出检索词本身，不要解释。

系统: {systemName} {systemDesc}
表: {tableChnName} ({tableName})
字段: {columnChnName} ({columnName})

检索词:"""


def rewrite_query(row):
    prompt = REWRITE_PROMPT.format(**{k: row.get(k, "") for k in ALL_FIELDS})
    text = _chat(prompt, max_tokens=64)
    return (text or row.get("columnChnName", "")).strip().splitlines()[0].strip()


# ---------------------------------------------------------------------------
# 判定（LLM）
# ---------------------------------------------------------------------------
JUDGE_PROMPT = """你是数据安全分类分级专家。请根据字段信息，从候选列表中选择最匹配的一个 category。
若都不匹配，category 返回 null。严格输出 JSON：{{"category": "...或null", "matched_catalog": "...或null", "reason": "50字内"}}

字段: 系统={systemName} 表={tableChnName}({tableName}) 字段={columnChnName}({columnName})

候选列表:
{candidates}
"""


def judge(row, candidates):
    cand_text = "\n".join(
        f"{i + 1}. category={c['category']} | asset={c['asset']}"
        for i, c in enumerate(candidates)
    ) or "(空)"
    fields = {k: row.get(k, "") for k in ALL_FIELDS}
    prompt = JUDGE_PROMPT.format(candidates=cand_text, **fields)
    text = _chat(prompt, max_tokens=256)
    return _parse_json_category(text)


def _parse_json_category(text):
    if not text:
        return None
    s = text.strip()
    start, end = s.find("{"), s.rfind("}")
    if start >= 0 and end > start:
        s = s[start : end + 1]
    try:
        obj = json.loads(s)
    except json.JSONDecodeError:
        return None
    cat = obj.get("category")
    return normalize_category(cat) if cat else None


def _chat(prompt, max_tokens=256):
    if not DASHSCOPE_API_KEY:
        raise SystemExit("缺少 DASHSCOPE_API_KEY，无法调用 LLM")
    payload = {
        "model": LLM_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0,
        "max_tokens": max_tokens,
    }
    headers = {"Authorization": f"Bearer {DASHSCOPE_API_KEY}"}
    resp = requests.post(
        f"{DASHSCOPE_BASE_URL}/chat/completions",
        headers=headers,
        json=payload,
        timeout=HTTP_TIMEOUT,
    )
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"]


# ---------------------------------------------------------------------------
# 策略：构造召回 query
# ---------------------------------------------------------------------------
def build_query(strategy, row):
    if strategy == "baseline":
        return row.get("columnChnName", "")
    if strategy == "multifield":
        parts = [row.get("tableChnName", ""), row.get("columnChnName", ""), row.get("columnName", "")]
        return " ".join(p for p in parts if p).strip()
    if strategy == "rewrite":
        return rewrite_query(row)
    raise ValueError(f"未知策略: {strategy}")


# ---------------------------------------------------------------------------
# 工具
# ---------------------------------------------------------------------------
def normalize_category(cat):
    if not cat:
        return ""
    return " ".join(str(cat).strip().split()).replace("｜", "|")


def hit_rank(gold, fused):
    gold = normalize_category(gold)
    for i, item in enumerate(fused):
        if item["category"] == gold:
            return i + 1
    return None


# ---------------------------------------------------------------------------
# 评测主流程
# ---------------------------------------------------------------------------
def run_strategy(strategy, rows, args):
    print(f"\n=== 策略: {strategy} ===")
    n = len(rows)
    rec_at_k = 0
    rec_vec = 0
    rec_bm25 = 0
    mrr = 0.0
    judge_total = 0
    judge_correct = 0
    e2e_correct = 0
    detail = []

    for idx, row in enumerate(rows):
        gold = normalize_category(row.get(GOLD_FIELD))
        try:
            q = build_query(strategy, row)
            vec = es_vector_search(q, args.topk, args.min_score)
            bm = es_bm25_search(q, args.topk)
        except requests.HTTPError as e:
            print(f"  [error] 第 {idx + 1} 行检索失败: {e}", file=sys.stderr)
            detail.append({"row": idx + 1, "error": str(e)})
            continue
        except Exception as e:  # noqa: BLE001
            print(f"  [error] 第 {idx + 1} 行: {e}", file=sys.stderr)
            detail.append({"row": idx + 1, "error": str(e)})
            continue

        fused = rrf_fuse(vec, bm)[: args.k]
        rank = hit_rank(gold, fused)
        vec_hit = hit_rank(gold, [{"category": normalize_category(x["category"]), "asset": x["asset"]} for x in vec])
        bm_hit = hit_rank(gold, [{"category": normalize_category(x["category"]), "asset": x["asset"]} for x in bm])

        if rank:
            rec_at_k += 1
            mrr += 1.0 / rank
        if vec_hit:
            rec_vec += 1
        if bm_hit:
            rec_bm25 += 1

        rec = {
            "row": idx + 1,
            "query": q,
            "gold": gold,
            "recall_rank": rank or "",
            "vec_hit": bool(vec_hit),
            "bm25_hit": bool(bm_hit),
            "top1": fused[0]["category"] if fused else "",
        }

        if args.judge:
            judge_total += 1
            chosen = judge(row, fused)
            rec["judged"] = chosen or "null"
            if chosen and chosen == gold:
                e2e_correct += 1
            # 判定准确率只在“召回命中”子集上统计
            if rank:
                if chosen and chosen == gold:
                    judge_correct += 1

        detail.append(rec)
        if (idx + 1) % 20 == 0:
            print(f"  进度 {idx + 1}/{n}")

    summary = OrderedDict()
    summary["strategy"] = strategy
    summary["samples"] = n
    summary[f"recall@{args.k}"] = _pct(rec_at_k, n)
    summary[f"vector_recall@{args.topk}"] = _pct(rec_vec, n)
    summary[f"bm25_recall@{args.topk}"] = _pct(rec_bm25, n)
    summary["MRR"] = round(mrr / n, 4) if n else 0
    if args.judge:
        summary["judge_acc(召回命中子集)"] = _pct(judge_correct, rec_at_k)
        summary["end2end_acc"] = _pct(e2e_correct, n)
    return summary, detail


def _pct(num, den):
    return f"{(100.0 * num / den):.1f}% ({num}/{den})" if den else "N/A"


def print_summary_table(summaries):
    print("\n" + "=" * 70)
    print("评测汇总")
    print("=" * 70)
    keys = list(summaries[0].keys())
    for s in summaries:
        print()
        for kk in keys:
            print(f"  {kk:<28}: {s.get(kk, '')}")


def write_reports(out_prefix, summaries, details_by_strategy):
    with open(f"{out_prefix}_summary.json", "w", encoding="utf-8") as f:
        json.dump(summaries, f, ensure_ascii=False, indent=2)
    for strat, detail in details_by_strategy.items():
        path = f"{out_prefix}_{strat}_detail.csv"
        if not detail:
            continue
        keys = sorted({k for d in detail for k in d.keys()})
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=keys)
            w.writeheader()
            w.writerows(detail)
    print(f"\n报告已写入: {out_prefix}_summary.json 及各策略明细 csv")


def main():
    parser = argparse.ArgumentParser(description="召回评测脚本")
    parser.add_argument("--data", required=True, help="标注数据文件 .csv/.xlsx")
    parser.add_argument("--strategy", default="all",
                        choices=["baseline", "multifield", "rewrite", "all"])
    parser.add_argument("--k", type=int, default=10, help="融合后 recall@K 的 K")
    parser.add_argument("--topk", type=int, default=10, help="每路检索 top-K")
    parser.add_argument("--min-score", type=float, default=0.0, dest="min_score",
                        help="向量检索 min_score（ES 归一化分），评测期建议设 0 先看全量召回")
    parser.add_argument("--judge", action="store_true", help="额外跑 LLM 判定，统计判定/端到端准确率")
    parser.add_argument("--limit", type=int, default=0, help="只评测前 N 条（0=全部）")
    parser.add_argument("--out", default="report", help="报告输出前缀")
    args = parser.parse_args()

    rows = load_data(args.data)
    if args.limit:
        rows = rows[: args.limit]
    if not rows:
        raise SystemExit("没有可评测的数据（检查文件与 expected_category 列）")
    print(f"加载样本 {len(rows)} 条；ES={ES_HOST}/{ES_INDEX}；embed={EMBED_MODEL}")

    strategies = ["baseline", "multifield", "rewrite"] if args.strategy == "all" else [args.strategy]
    summaries = []
    details = {}
    t0 = time.time()
    for strat in strategies:
        s, d = run_strategy(strat, rows, args)
        summaries.append(s)
        details[strat] = d

    print_summary_table(summaries)
    write_reports(args.out, summaries, details)
    print(f"\n耗时 {time.time() - t0:.1f}s")


if __name__ == "__main__":
    main()
