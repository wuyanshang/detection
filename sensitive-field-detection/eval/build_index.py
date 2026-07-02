#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
关键项目录索引构建脚本 (build-index)

用途：把"分类分级目录"Excel（一级/二级/三级/四级 + 各级定义 + 内容 + 最低安全级别）
构建成 Elasticsearch 索引，支撑安全分类分级的向量 + BM25 混合召回。

相比旧索引（只 embed 了叶子名 asset）的关键改进：
  1. content（枚举示例）、level2/level3 定义 全部入库；
  2. 向量 embedding 文本 = 四级名 + 三级定义 + content（信号最全）；
  3. content 里的枚举示例（如"个人姓名/性别/证件号码"）拆成 example 文档，
     每条指回同一 category —— 输入"客户姓名"可直接命中"个人姓名"；
  4. BM25 检索字段扩展为 asset + content + example。

对应设计文档：../安全分类分级-召回优化重构方案.md 第 3.1 节。

Excel 列（表头含关键词即可，支持合并单元格：一级~三级及其定义向下填充）：
  一级子类 / 二级子类 / 二级子类定义说明 / 三级子类 / 三级子类定义说明 /
  四级子类 / 内容 / 最低安全级别参考

环境变量：
  ES_HOST            例 http://127.0.0.1:9200
  ES_USERNAME        可选
  ES_PASSWORD        可选
  ES_INDEX           默认 safe_all_topic_v2（新索引，灰度验证后再切换生产 index）
  ES_ANALYZER        中文分词器，默认 ik_max_word；若未装 IK 插件请设为 standard
  DASHSCOPE_API_KEY  生成 embedding 必需
  DASHSCOPE_BASE_URL 默认 https://dashscope.aliyuncs.com/compatible-mode/v1
  EMBED_MODEL        默认 text-embedding-v4
  EMBED_DIM          默认 1024

用法：
  # 重建索引（删除旧的同名索引后重建 mapping，再灌数据）
  python build_index.py --data catalog.xlsx --recreate

  # 不建示例文档，只建 category 文档
  python build_index.py --data catalog.xlsx --recreate --no-examples

  # 只解析打印前 5 条，不写 ES（自检列映射/拆分是否正确）
  python build_index.py --data catalog.xlsx --dry-run --limit 5
"""
import argparse
import json
import os
import re
import sys
import time

import requests

# ---------------------------------------------------------------------------
# 配置
# ---------------------------------------------------------------------------
ES_HOST = os.getenv("ES_HOST", "http://127.0.0.1:9200").rstrip("/")
ES_USERNAME = os.getenv("ES_USERNAME") or None
ES_PASSWORD = os.getenv("ES_PASSWORD") or None
ES_INDEX = os.getenv("ES_INDEX", "safe_all_topic_v2")
ES_ANALYZER = os.getenv("ES_ANALYZER", "ik_max_word")

DASHSCOPE_API_KEY = os.getenv("DASHSCOPE_API_KEY")
DASHSCOPE_BASE_URL = os.getenv(
    "DASHSCOPE_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1"
).rstrip("/")
EMBED_MODEL = os.getenv("EMBED_MODEL", "text-embedding-v4")
EMBED_DIM = int(os.getenv("EMBED_DIM", "1024"))

HTTP_TIMEOUT = int(os.getenv("HTTP_TIMEOUT", "60"))
EMBED_BATCH = int(os.getenv("EMBED_BATCH", "10"))

_es = requests.Session()
if ES_USERNAME and ES_PASSWORD:
    _es.auth = (ES_USERNAME, ES_PASSWORD)


# ---------------------------------------------------------------------------
# 1. 读取 Excel + 合并单元格向下填充
# ---------------------------------------------------------------------------
# 列角色：按表头关键词识别，先判定"定义/说明"列，避免与层级名列混淆
def resolve_columns(header):
    roles = {}
    for i, raw in enumerate(header):
        h = (raw or "").replace(" ", "")
        if not h:
            continue
        is_def = ("定义" in h) or ("说明" in h)
        if "一级" in h:
            roles[i] = "l1_def" if is_def else "l1"
        elif "二级" in h:
            roles[i] = "l2_def" if is_def else "l2"
        elif "三级" in h:
            roles[i] = "l3_def" if is_def else "l3"
        elif "四级" in h:
            roles[i] = "l4"
        elif "内容" in h:
            roles[i] = "content"
        elif ("安全" in h) and (("级别" in h) or ("等级" in h)):
            roles[i] = "security_level"
    return roles


def load_catalog(path):
    """读取 xlsx，返回规范化后的行列表（合并单元格已向下填充）。"""
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".xls"):
        raise SystemExit(f"仅支持 .xlsx/.xls，当前：{ext}")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("需要 openpyxl：pip install openpyxl")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    roles = resolve_columns(header)
    if "l4" not in roles.values() or "content" not in roles.values():
        raise SystemExit(f"未识别到【四级子类】或【内容】列，识别结果={roles}，表头={header}")

    # 需要向下填充的层级列（合并单元格只在首行有值）
    FILL_ROLES = {"l1", "l1_def", "l2", "l2_def", "l3", "l3_def"}
    carry = {}
    out = []
    for r in rows_iter:
        rec = {}
        for i, role in roles.items():
            val = norm(r[i]) if i < len(r) else ""
            if role in FILL_ROLES:
                if val:
                    carry[role] = val
                val = carry.get(role, "")
            rec[role] = val
        # 四级名与内容都空的行跳过（空行/小计行）
        if not rec.get("l4") and not rec.get("content"):
            continue
        out.append(rec)
    return out


def norm(s):
    """去除 CJK 文本里的换行/多余空白（OCR/换行产生的"个人 姓名"→"个人姓名"）。"""
    if s is None:
        return ""
    t = str(s).replace("\u3000", " ").strip()
    # 折叠所有空白：本目录为中文文本，去空白不影响 FATCA 等连续英文
    t = re.sub(r"\s+", "", t)
    return t


# ---------------------------------------------------------------------------
# 2. content 拆分为枚举示例
# ---------------------------------------------------------------------------
_EXAMPLE_SPLIT = re.compile(r"[、,，;；]")


def extract_examples(content):
    """从"指XX数据，如 A、B、C 等。"中抽出 [A, B, C]。抽不到返回 []。"""
    if not content:
        return []
    m = re.search(r"如(.+?)等", content)
    seg = m.group(1) if m else None
    if not seg:
        return []
    items = []
    for it in _EXAMPLE_SPLIT.split(seg):
        it = it.strip().strip("。.．:：")
        it = re.sub(r"^(个人|其|该)", "", it)  # 去掉"个人收入状况"类前缀噪声(保守)
        if 2 <= len(it) <= 25 and it not in items:
            items.append(it)
    return items


# ---------------------------------------------------------------------------
# 3. 构建文档
# ---------------------------------------------------------------------------
def build_docs(rows, with_examples=True):
    """把行转成 ES 文档（category 文档 + 可选 example 文档）。"""
    docs = []
    seen_cat = set()
    for rec in rows:
        levels = [rec.get("l1", ""), rec.get("l2", ""), rec.get("l3", ""), rec.get("l4", "")]
        levels = [x for x in levels if x]
        if not levels:
            continue
        asset = rec.get("l4") or levels[-1]
        category = "|".join(levels)
        if category in seen_cat:
            print(f"  [warn] 重复 category，跳过：{category}", file=sys.stderr)
            continue
        seen_cat.add(category)

        content = rec.get("content", "")
        l2_def = rec.get("l2_def", "")
        l3_def = rec.get("l3_def", "")
        sec = parse_int(rec.get("security_level"))
        topic = rec.get("l1", "")

        # 向量文本：叶子名 + 三级定义 + content（信号最全）
        embed_text = "。".join(x for x in [asset, l3_def, content] if x)

        cat_doc = {
            "_id": f"cat::{category}",
            "doc_type": "category",
            "topic": topic,
            "level1": rec.get("l1", ""),
            "level2": rec.get("l2", ""),
            "level3": rec.get("l3", ""),
            "asset": asset,
            "category": category,
            "level2_definition": l2_def,
            "level3_definition": l3_def,
            "content": content,
            "security_level": sec,
            "_embed_text": embed_text,
        }
        docs.append(cat_doc)

        if with_examples:
            for i, ex in enumerate(extract_examples(content)):
                docs.append({
                    "_id": f"ex::{category}::{i}",
                    "doc_type": "example",
                    "topic": topic,
                    "asset": asset,
                    "category": category,
                    "example": ex,
                    "content": content,
                    "security_level": sec,
                    "_embed_text": ex,  # 示例项本身做向量，利于命中"客户姓名→个人姓名"
                })
    return docs


def parse_int(s):
    if s is None or s == "":
        return None
    m = re.search(r"\d+", str(s))
    return int(m.group()) if m else None


# ---------------------------------------------------------------------------
# 4. Embedding（批量）
# ---------------------------------------------------------------------------
def embed_batch(texts):
    if not DASHSCOPE_API_KEY:
        raise SystemExit("缺少 DASHSCOPE_API_KEY，无法生成 embedding")
    payload = {"model": EMBED_MODEL, "input": texts, "dimensions": EMBED_DIM}
    headers = {"Authorization": f"Bearer {DASHSCOPE_API_KEY}"}
    resp = requests.post(
        f"{DASHSCOPE_BASE_URL}/embeddings",
        headers=headers, json=payload, timeout=HTTP_TIMEOUT,
    )
    resp.raise_for_status()
    data = sorted(resp.json()["data"], key=lambda d: d.get("index", 0))
    return [d["embedding"] for d in data]


def attach_embeddings(docs):
    total = len(docs)
    for start in range(0, total, EMBED_BATCH):
        batch = docs[start:start + EMBED_BATCH]
        vecs = embed_batch([d["_embed_text"] for d in batch])
        for d, v in zip(batch, vecs):
            d["asset_embedding"] = v
        print(f"  embedding 进度 {min(start + EMBED_BATCH, total)}/{total}")
    return docs


# ---------------------------------------------------------------------------
# 5. 索引 mapping 与写入
# ---------------------------------------------------------------------------
def index_mapping():
    text_zh = {"type": "text", "analyzer": ES_ANALYZER}
    return {
        "settings": {"number_of_shards": 1, "number_of_replicas": 0},
        "mappings": {
            "properties": {
                "doc_type": {"type": "keyword"},
                "topic": {"type": "keyword"},
                "level1": {"type": "keyword"},
                "level2": {"type": "keyword"},
                "level3": {"type": "keyword"},
                "asset": {**text_zh, "fields": {"kw": {"type": "keyword"}}},
                "category": {"type": "keyword"},
                "example": text_zh,
                "content": text_zh,
                "level2_definition": text_zh,
                "level3_definition": text_zh,
                "security_level": {"type": "integer"},
                "asset_embedding": {
                    "type": "dense_vector",
                    "dims": EMBED_DIM,
                    "index": True,
                    "similarity": "cosine",
                },
            }
        },
    }


def recreate_index():
    _es.delete(f"{ES_HOST}/{ES_INDEX}", timeout=HTTP_TIMEOUT)  # 忽略 404
    resp = _es.put(
        f"{ES_HOST}/{ES_INDEX}",
        json=index_mapping(),
        headers={"Content-Type": "application/json"},
        timeout=HTTP_TIMEOUT,
    )
    if resp.status_code >= 300:
        raise SystemExit(f"创建索引失败({resp.status_code}): {resp.text}\n"
                         f"若报 analyzer [{ES_ANALYZER}] 不存在，请安装 IK 插件或设 ES_ANALYZER=standard")
    print(f"索引已重建：{ES_INDEX}（analyzer={ES_ANALYZER}, dims={EMBED_DIM}）")


def bulk_index(docs):
    lines = []
    for d in docs:
        _id = d.pop("_id")
        d.pop("_embed_text", None)
        lines.append(json.dumps({"index": {"_index": ES_INDEX, "_id": _id}}, ensure_ascii=False))
        lines.append(json.dumps(d, ensure_ascii=False))
    body = "\n".join(lines) + "\n"
    resp = _es.post(
        f"{ES_HOST}/_bulk",
        data=body.encode("utf-8"),
        headers={"Content-Type": "application/x-ndjson"},
        timeout=HTTP_TIMEOUT,
    )
    resp.raise_for_status()
    result = resp.json()
    if result.get("errors"):
        errs = [it for it in result["items"] if it.get("index", {}).get("status", 200) >= 300]
        print(f"  [warn] 写入部分失败 {len(errs)} 条，示例：{errs[:2]}", file=sys.stderr)
    return len(docs)


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="关键项目录索引构建")
    ap.add_argument("--data", required=True, help="目录 Excel 文件 .xlsx")
    ap.add_argument("--recreate", action="store_true", help="删除并重建索引 mapping")
    ap.add_argument("--no-examples", action="store_true", help="不生成枚举示例文档")
    ap.add_argument("--dry-run", action="store_true", help="只解析打印，不调用 embedding/不写 ES")
    ap.add_argument("--limit", type=int, default=0, help="只处理前 N 行（自检用）")
    args = ap.parse_args()

    rows = load_catalog(args.data)
    if args.limit:
        rows = rows[: args.limit]
    print(f"读取目录行 {len(rows)} 条；ES={ES_HOST}/{ES_INDEX}")

    docs = build_docs(rows, with_examples=not args.no_examples)
    n_cat = sum(1 for d in docs if d["doc_type"] == "category")
    n_ex = sum(1 for d in docs if d["doc_type"] == "example")
    print(f"生成文档 {len(docs)} 个（category={n_cat}, example={n_ex}）")

    if args.dry_run:
        for d in docs[:20]:
            preview = {k: v for k, v in d.items() if k != "_embed_text"}
            print(json.dumps(preview, ensure_ascii=False))
            print(f"    _embed_text = {d['_embed_text'][:60]}...")
        print("\n[dry-run] 未调用 embedding、未写 ES。")
        return

    t0 = time.time()
    attach_embeddings(docs)
    if args.recreate:
        recreate_index()
    n = bulk_index(docs)
    _es.post(f"{ES_HOST}/{ES_INDEX}/_refresh", timeout=HTTP_TIMEOUT)
    print(f"\n完成：写入 {n} 个文档，耗时 {time.time() - t0:.1f}s")


if __name__ == "__main__":
    main()
